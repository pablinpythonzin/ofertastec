"""
═══════════════════════════════════════════════════════════════════════
  GERAR PRODUTOS — OfertasTec
  Cruza as fotos de produtos/ com a planilha e gera o código pro site
═══════════════════════════════════════════════════════════════════════

O QUE FAZ:
  1. Você escolhe uma categoria
  2. Olha as fotos .jpg que existem em produtos/
  3. Cruza cada foto com a planilha (pega nome + link de afiliado)
  4. PULA os produtos vermelhos (link quebrado) automaticamente
  5. INCLUI verde + amarelo (regra do OfertasTec)
  6. Gera o código pronto pra colar no index.html
  7. Salva o código num arquivo .txt também (codigo_gerado.txt)

REGRA DE CORES:
  🟢 Verde   = OK          → ENTRA
  🟡 Amarelo = pouco estoque → ENTRA (reabastece sozinho)
  🔴 Vermelho = quebrado    → PULA (link quebrado espanta cliente)
  🟣 Roxo    = duplicado    → PULA

COMO USAR:
  1. Salva na pasta OfertasTec/
  2. Processa as fotos antes (com processar_fotos.py)
  3. Roda: python gerar_produtos.py
  4. Escolhe a categoria
  5. Copia o código gerado e cola no index.html (no lugar da categoria)

COMO COLAR NO SITE:
  - Abra o index.html
  - Ache a linha que começa com {id:'nome-da-categoria',
  - Substitua o bloco inteiro daquela categoria (de { até }, )
    pelo código que esse script gerou
═══════════════════════════════════════════════════════════════════════
"""

import sys
import unicodedata
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("\n❌ ERRO: 'openpyxl' não instalado. Rode: pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════

PASTA_PRODUTOS = "produtos"
ARQUIVO_PLANILHA = "Produtos_ml.xlsx"
ABA_PLANILHA = "Produtos Renomear"
ARQUIVO_SAIDA = "codigo_gerado.txt"

COL_NUM = 2          # B - #
COL_CATEGORIA = 3    # C - Categoria
COL_LINK = 4         # D - Link
COL_NOME = 5         # E - Nome

# Cores que PULAM
COR_VERMELHO = "FFC00000"
COR_ROXO = "FF7030A0"

# Mapa: nome-da-planilha → dados pro site (slug, nome-exibição, icone, cor, busca-ML)
# As 30 categorias oficiais
CATEGORIAS = {
    "Mouse sem fio":                  ("mouse-sem-fio", "Mouse Sem Fio", "ti-mouse", "#38BFFF", "mouse sem fio bluetooth"),
    "Mouse com fio":                  ("mouse-com-fio", "Mouse Com Fio", "ti-mouse", "#2ED8BC", "mouse com fio usb"),
    "Mouse gamer sem fio":            ("mouse-gamer-sem-fio", "Mouse Gamer Sem Fio", "ti-mouse", "#FF3D7F", "mouse gamer sem fio rgb"),
    "Mouse gamer com fio":            ("mouse-gamer-com-fio", "Mouse Gamer Com Fio", "ti-mouse", "#AA7AFF", "mouse gamer com fio rgb"),
    "Mouse vertical (ergonômico)":    ("mouse-vertical", "Mouse Vertical", "ti-mouse", "#9C6ADE", "mouse vertical ergonomico"),
    "Teclado com fio":                ("teclado-com-fio", "Teclado Com Fio", "ti-keyboard", "#FFE000", "teclado com fio usb"),
    "Teclado sem fio":                ("teclado-sem-fio", "Teclado Sem Fio", "ti-keyboard", "#FFA040", "teclado sem fio bluetooth"),
    "Teclado gamer com fio":          ("teclado-gamer-com-fio", "Teclado Gamer Com Fio", "ti-keyboard", "#FF6B35", "teclado gamer mecanico rgb"),
    "Teclado gamer sem fio":          ("teclado-gamer-sem-fio", "Teclado Gamer Sem Fio", "ti-keyboard", "#FF4D6D", "teclado gamer sem fio mecanico"),
    "Fone sem fio (bluetooth)":       ("fone-sem-fio", "Fone Sem Fio Bluetooth", "ti-headphones", "#E040FB", "fone bluetooth sem fio in ear"),
    "Fone com fio":                   ("fone-com-fio", "Fone Com Fio", "ti-headphones", "#00E5FF", "fone com fio p2 ouvido"),
    "Fone gamer (headset sem fio)":   ("fone-gamer-sem-fio", "Headset Gamer Sem Fio", "ti-headset", "#52FF8E", "headset gamer sem fio"),
    "Headset Gamer Com Fio":          ("fone-gamer-com-fio", "Headset Gamer Com Fio", "ti-headset", "#76FF03", "headset gamer com fio"),
    "Caixa de som bluetooth":         ("caixa-som-bluetooth", "Caixa de Som Bluetooth", "ti-speaker", "#FF8C42", "caixa de som bluetooth portatil"),
    "Caixa de som com fio":           ("caixa-som-com-fio", "Caixa de Som Com Fio", "ti-speaker", "#FFA552", "caixa de som com fio usb"),
    "Cabo USB-C":                     ("cabo-usbc", "Cabo USB-C", "ti-usb", "#FFEA00", "cabo usb tipo c carregador"),
    "Cabo Lightning (iPhone)":        ("cabo-lightning", "Cabo Lightning (iPhone)", "ti-plug-connected", "#C0C0C0", "cabo lightning iphone"),
    "Power bank":                     ("power-bank", "Power Bank", "ti-battery-charging", "#4CD964", "power bank carregador portatil"),
    "Suporte p/ celular (carro)":     ("suporte-celular-carro", "Suporte de Celular (Carro)", "ti-device-mobile", "#5AC8FA", "suporte celular carro"),
    "Película protetora":             ("pelicula-protetora", "Pelicula Protetora", "ti-shield", "#34AADC", "pelicula protetora celular"),
    "Capinha celular":                ("capinha-celular", "Capinha de Celular", "ti-device-mobile", "#FF2D55", "capa capinha celular"),
    "Mousepad simples":               ("mousepad-simples", "Mousepad Simples", "ti-square-rotated", "#8E8E93", "mousepad simples"),
    "Mousepad gamer (grande)":        ("mousepad-gamer", "Mousepad Gamer (Grande)", "ti-square-rotated", "#AF52DE", "mousepad gamer grande rgb"),
    "Controle PC (sem fio)":          ("controle-pc-sem-fio", "Controle PC (Sem Fio)", "ti-device-gamepad-2", "#FF6482", "controle pc sem fio"),
    "Controle PC (com fio)":          ("controle-pc-com-fio", "Controle PC (Com Fio)", "ti-device-gamepad-2", "#FF9500", "controle pc com fio usb"),
    "Controle para celular":          ("controle-celular", "Controle para Celular", "ti-device-gamepad", "#30D158", "controle gamepad celular"),
    "Adaptador de tomada":            ("adaptador-tomada", "Adaptador de Tomada", "ti-plug", "#FFD60A", "adaptador tomada"),
    "Cabo HDMI":                      ("cabo-hdmi", "Cabo HDMI", "ti-cable", "#64D2FF", "cabo hdmi 4k"),
    "Pen drive":                      ("pen-drive", "Pen Drive", "ti-usb", "#5E5CE6", "pen drive usb"),
    "HD externo":                     ("hd-externo", "HD Externo", "ti-device-sd-card", "#BF5AF2", "hd externo portatil"),
}


def remover_acentos(texto):
    if not texto:
        return ""
    nfkd = unicodedata.normalize('NFKD', str(texto))
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def gerar_slug(nome):
    nome = remover_acentos(nome).lower()
    nome = re.sub(r'[^a-z0-9]+', '-', nome)
    return re.sub(r'-+', '-', nome).strip('-')


def cor_linha(ws, row):
    cell = ws.cell(row=row, column=COL_NOME)
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = cell.fill.fgColor.rgb
        return None if rgb in ('00000000', None) else rgb
    return None


def carregar_planilha():
    if not Path(ARQUIVO_PLANILHA).exists():
        print(f"\n❌ Planilha '{ARQUIVO_PLANILHA}' não encontrada!")
        return None
    try:
        wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
    except PermissionError:
        print("\n❌ Planilha ABERTA no Excel! Feche e tente de novo.")
        return None
    ws = wb[ABA_PLANILHA]

    # produtos[categoria] = lista de dicts
    produtos = {}
    for row in range(2, ws.max_row + 1):
        cat = ws.cell(row=row, column=COL_CATEGORIA).value
        link = ws.cell(row=row, column=COL_LINK).value
        nome = ws.cell(row=row, column=COL_NOME).value
        num = ws.cell(row=row, column=COL_NUM).value
        if cat and link and nome and str(link).strip().startswith('http'):
            cat = str(cat).strip()
            produtos.setdefault(cat, []).append({
                'num': num,
                'link': str(link).strip(),
                'nome': str(nome).strip(),
                'slug': gerar_slug(str(nome).strip()),
                'cor': cor_linha(ws, row),
            })
    return produtos


def escolher_categoria():
    print("\n  CATEGORIAS DISPONÍVEIS:\n")
    cats = list(CATEGORIAS.keys())
    for i, c in enumerate(cats, 1):
        print(f"  [{i:>2}] {c}")
    print("  [0] Sair")
    while True:
        r = input("\n  Escolha a categoria: ").strip()
        try:
            n = int(r)
            if n == 0:
                return None
            if 1 <= n <= len(cats):
                return cats[n - 1]
        except ValueError:
            pass
        print("  ⚠️ Número inválido.")


def main():
    print("\n" + "=" * 65)
    print("  🏗️  GERAR PRODUTOS — OfertasTec")
    print("=" * 65)

    pasta_produtos = Path.cwd() / PASTA_PRODUTOS
    if not pasta_produtos.exists():
        print(f"\n❌ Pasta '{PASTA_PRODUTOS}/' não encontrada!")
        input("\nENTER..."); return

    print("\n📊 Carregando planilha...")
    produtos_planilha = carregar_planilha()
    if produtos_planilha is None:
        input("\nENTER..."); return
    print("✅ Planilha carregada")

    # fotos disponíveis (jpg na raiz de produtos/)
    fotos = {f.name for f in pasta_produtos.iterdir()
             if f.is_file() and f.suffix.lower() == '.jpg'}
    print(f"✅ {len(fotos)} fotos .jpg em produtos/")

    while True:
        cat_planilha = escolher_categoria()
        if cat_planilha is None:
            break

        slug, nome_exib, icone, cor, busca = CATEGORIAS[cat_planilha]
        lista = produtos_planilha.get(cat_planilha, [])

        if not lista:
            print(f"\n  ⚠️ Nenhum produto de '{cat_planilha}' na planilha.")
            input("\n  ENTER..."); continue

        # cruzar com fotos + aplicar regra de cores
        entram = []      # (link, nome, foto)
        sem_foto = []    # (link, nome)  -> verde/amarelo mas sem foto
        pulados = []     # (nome, motivo)
        for p in lista:
            if p['cor'] == COR_VERMELHO:
                pulados.append((p['nome'], "🔴 vermelho (quebrado)"))
                continue
            if p['cor'] == COR_ROXO:
                pulados.append((p['nome'], "🟣 roxo (duplicado)"))
                continue
            # verde ou amarelo → entra
            foto = p['slug'] + '.jpg'
            if foto in fotos:
                entram.append((p['link'], p['nome'], foto))
            else:
                sem_foto.append((p['link'], p['nome']))

        # relatório
        print("\n" + "=" * 65)
        print(f"  📦 {nome_exib}")
        print("=" * 65)
        print(f"  ✅ {len(entram)} com foto")
        print(f"  📝 {len(sem_foto)} sem foto (entram só com link)")
        print(f"  ⏭️  {len(pulados)} pulados (vermelho/roxo)")
        if pulados:
            print("\n  Pulados:")
            for nome, motivo in pulados:
                print(f"     {motivo} | {nome[:50]}")

        # montar o código
        linhas = []
        for link, nome, foto in entram:
            ne = nome.replace("'", "\\'")
            linhas.append(f"     {{u:'{link}', n:'{ne}', img:'produtos/{foto}'}}")
        for link, nome in sem_foto:
            ne = nome.replace("'", "\\'")
            linhas.append(f"     {{u:'{link}', n:'{ne}'}}")

        if not linhas:
            print("\n  ❌ Nenhum produto pra gerar (só tinha vermelho/roxo).")
            input("\n  ENTER..."); continue

        bloco = (f"  {{id:'{slug}', nome:'{nome_exib}', icon:'{icone}', cor:'{cor}', mlSearch:'{busca}',\n"
                 f"   produtos:[\n" + ",\n".join(linhas) + "\n   ]},")

        # mostrar na tela
        print("\n" + "=" * 65)
        print("  📋 CÓDIGO GERADO (copie e cole no index.html):")
        print("=" * 65 + "\n")
        print(bloco)
        print("\n" + "=" * 65)

        # salvar no txt
        with open(ARQUIVO_SAIDA, 'w', encoding='utf-8') as f:
            f.write(bloco + "\n")
        print(f"  💾 Também salvo em: {ARQUIVO_SAIDA}")
        print("\n  📝 Como colar:")
        print(f"     1. Abra o index.html")
        print(f"     2. Ache a linha:  {{id:'{slug}',")
        print(f"     3. Substitua o bloco daquela categoria pelo código acima")
        print(f"        (do {{ até o }}, do fim)")
        input("\n  ENTER pra voltar ao menu...")

    print("\n  👋 Até a próxima!\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n👋 Saindo...")
    except Exception as e:
        print(f"\n❌ ERRO: {e}")
        input("\nENTER pra sair...")
