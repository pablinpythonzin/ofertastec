"""
═══════════════════════════════════════════════════════════════════════
  VERIFICADOR DE LINKS — OfertasTec
  Abre cada link da planilha, você confirma se está OK ou quebrado
═══════════════════════════════════════════════════════════════════════

COMO USAR:
1. Salva esse arquivo na pasta OfertasTec/
2. Feche o Excel (importante!)
3. Roda no PowerShell:
     python verificar_links.py

O QUE FAZ:
- Abre cada link da planilha no navegador
- Você olha a página com seus próprios olhos
- Responde: tá OK, tá com pouco estoque, ou tá quebrado
- Script marca a planilha automaticamente:
    🟢 Verde   = OK
    🟡 Amarelo = Pouco estoque (<10 unidades) ou pouco confiável
    🔴 Vermelho = Quebrado / vitrine vazia / fora do ar

ATALHOS:
  s = link OK (verde)
  a = pouco estoque ou suspeito (amarelo)
  n = link quebrado (vermelho)
  p = pular (não marca nada)
  v = abrir o link de novo no navegador
  q = sair e salvar

FUNCIONALIDADES BONUS:
- Retoma de onde parou (não verifica links já marcados)
- Filtra por categoria (verifica só uma categoria de cada vez)
- Mostra progresso (37/130)
- Salva planilha automaticamente após cada verificação

═══════════════════════════════════════════════════════════════════════
"""

import os
import sys
import webbrowser
import time
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ImportError:
    print("\n❌ ERRO: 'openpyxl' não instalado.")
    print("Rode: pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# CONFIGURAÇÕES
# ═══════════════════════════════════════════════════════════════════════

ARQUIVO_PLANILHA = "Produtos_ml.xlsx"
ABA_PLANILHA = "Produtos Renomear"

COL_NUM = 2          # B - #
COL_CATEGORIA = 3    # C - Categoria
COL_LINK = 4         # D - Link (clique pra abrir)
COL_NOME = 5         # E - Nome do Produto
COL_PRECO = 6        # F - Preço
COL_MARCA = 7        # G - Marca
COL_OBS = 8          # H - Observações

# Cores (ARGB)
COR_VERDE = "FF92D050"     # OK
COR_AMARELO = "FFFFFF00"   # Atenção
COR_VERMELHO = "FFC00000"  # Quebrado
COR_BRANCO = None          # Sem cor (resetar)


# ═══════════════════════════════════════════════════════════════════════
# FUNÇÕES
# ═══════════════════════════════════════════════════════════════════════

def pintar_linha(ws, row, cor_hex):
    """Pinta a linha inteira (colunas B-H) com a cor."""
    if cor_hex is None:
        fill = PatternFill(fill_type=None)
    else:
        fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
    for col in range(COL_NUM, COL_OBS + 1):
        ws.cell(row=row, column=col).fill = fill


def cor_atual(ws, row):
    """Detecta a cor atual da linha (pela coluna do nome)."""
    cell = ws.cell(row=row, column=COL_NOME)
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = cell.fill.fgColor.rgb
        if rgb in ('00000000', None):
            return None
        return rgb
    return None


def label_cor(rgb):
    if rgb is None:
        return "⚪ não verificada"
    elif rgb == COR_VERDE:
        return "🟢 OK"
    elif rgb == COR_AMARELO:
        return "🟡 atenção"
    elif rgb == COR_VERMELHO:
        return "🔴 quebrada"
    else:
        return f"❓ ({rgb})"


def carregar_produtos(ws):
    """Carrega todos os produtos com link válido."""
    produtos = []
    for row in range(2, ws.max_row + 1):
        num = ws.cell(row=row, column=COL_NUM).value
        cat = ws.cell(row=row, column=COL_CATEGORIA).value
        link = ws.cell(row=row, column=COL_LINK).value
        nome = ws.cell(row=row, column=COL_NOME).value

        if num and link and str(link).strip().startswith('http'):
            produtos.append({
                'row': row,
                'num': num,
                'cat': str(cat).strip() if cat else '',
                'link': str(link).strip(),
                'nome': str(nome).strip() if nome else '',
                'cor': cor_atual(ws, row),
            })
    return produtos


def menu_filtro(produtos):
    """Menu pra escolher filtros."""
    # contar por categoria
    por_cat = {}
    for p in produtos:
        por_cat.setdefault(p['cat'], []).append(p)

    # contar status
    todos = len(produtos)
    n_verde = sum(1 for p in produtos if p['cor'] == COR_VERDE)
    n_amarelo = sum(1 for p in produtos if p['cor'] == COR_AMARELO)
    n_vermelho = sum(1 for p in produtos if p['cor'] == COR_VERMELHO)
    n_branco = sum(1 for p in produtos if p['cor'] is None)

    print("\n" + "=" * 65)
    print("  🔍 VERIFICADOR DE LINKS — Status atual")
    print("=" * 65)
    print(f"\n  Total: {todos} produtos com link")
    print(f"  🟢 OK: {n_verde}")
    print(f"  🟡 Atenção: {n_amarelo}")
    print(f"  🔴 Quebrados: {n_vermelho}")
    print(f"  ⚪ Não verificados: {n_branco}")

    print("\n  ESCOLHE O QUE FAZER:")
    print(f"  [1] Verificar SÓ NÃO VERIFICADOS ({n_branco})")
    print(f"  [2] Verificar uma CATEGORIA específica")
    print(f"  [3] RE-verificar todos (incluso os já marcados)")
    print(f"  [4] Verificar só os AMARELOS ({n_amarelo})")
    print(f"  [0] Sair")

    while True:
        op = input("\n  Escolha [0-4]: ").strip()
        if op == '0':
            return None
        if op == '1':
            return [p for p in produtos if p['cor'] is None]
        if op == '2':
            return escolher_categoria(por_cat)
        if op == '3':
            return produtos
        if op == '4':
            return [p for p in produtos if p['cor'] == COR_AMARELO]
        print("  ⚠️ Opção inválida.")


def escolher_categoria(por_cat):
    print("\n  Categorias:")
    cats = sorted(por_cat.keys())
    for i, c in enumerate(cats, 1):
        total = len(por_cat[c])
        branco = sum(1 for p in por_cat[c] if p['cor'] is None)
        print(f"  [{i:>2}] {c:35} ({total} produtos, {branco} não verificados)")

    while True:
        op = input("\n  Escolha categoria [n° ou 0 pra voltar]: ").strip()
        if op == '0':
            return []
        try:
            i = int(op)
            if 1 <= i <= len(cats):
                cat = cats[i - 1]
                # filtrar só não verificados dessa categoria?
                subop = input("  Só os não verificados? (s/n): ").strip().lower()
                if subop == 's':
                    return [p for p in por_cat[cat] if p['cor'] is None]
                return por_cat[cat]
        except ValueError:
            pass
        print("  ⚠️ Opção inválida.")


def processar_lista(ws, wb, produtos):
    """Processa cada produto da lista."""
    total = len(produtos)
    if total == 0:
        print("\n  ✅ Nenhum produto pra verificar nessa seleção!")
        input("\n  Pressione ENTER...")
        return

    print(f"\n  🚀 Processando {total} link(s)\n")

    verde = 0
    amarelo = 0
    vermelho = 0
    puladas = 0

    for i, p in enumerate(produtos, 1):
        print("\n" + "─" * 65)
        print(f"  📌 LINK {i}/{total}")
        print("─" * 65)
        print(f"  #{p['num']} | {p['cat']}")
        print(f"  Nome: {p['nome'][:60] if p['nome'] else '(sem nome)'}")
        print(f"  Link: {p['link']}")
        print(f"  Status atual: {label_cor(p['cor'])}")
        print("\n  🌐 Abrindo no navegador...")

        webbrowser.open(p['link'])
        time.sleep(1)  # pequena espera pra abrir

        print("\n  AVALIE A PÁGINA:")
        print("    s = 🟢 OK (produto disponível, bom estoque)")
        print("    a = 🟡 ATENÇÃO (estoque baixo <10 unidades, ou suspeito)")
        print("    n = 🔴 QUEBRADO (vitrine vazia, 404, fora do ar)")
        print("    p = pular (não muda nada)")
        print("    v = abrir o link de novo")
        print("    q = sair e salvar")

        while True:
            resp = input("\n  → Resposta: ").strip().lower()

            if resp == 'q':
                print("\n  💾 Salvando planilha...")
                _salvar(wb)
                _resumo(verde, amarelo, vermelho, puladas, i - 1)
                input("\n  Pressione ENTER pra voltar ao menu...")
                return

            if resp == 'p':
                puladas += 1
                break

            if resp == 'v':
                webbrowser.open(p['link'])
                continue

            if resp == 's':
                pintar_linha(ws, p['row'], COR_VERDE)
                verde += 1
                print("  ✅ Marcado VERDE (OK)")
                _salvar(wb)
                break

            if resp == 'a':
                pintar_linha(ws, p['row'], COR_AMARELO)
                amarelo += 1
                # opcional: anotação
                obs_atual = ws.cell(row=p['row'], column=COL_OBS).value or ''
                if 'estoque baixo' not in obs_atual.lower():
                    nova_obs = (obs_atual + ' | estoque baixo').strip(' |')
                    ws.cell(row=p['row'], column=COL_OBS).value = nova_obs
                print("  🟡 Marcado AMARELO (atenção - estoque baixo)")
                _salvar(wb)
                break

            if resp == 'n':
                pintar_linha(ws, p['row'], COR_VERMELHO)
                vermelho += 1
                obs_atual = ws.cell(row=p['row'], column=COL_OBS).value or ''
                if 'fora do ar' not in obs_atual.lower():
                    nova_obs = (obs_atual + ' | fora do ar').strip(' |')
                    ws.cell(row=p['row'], column=COL_OBS).value = nova_obs
                print("  🔴 Marcado VERMELHO (quebrado)")
                _salvar(wb)
                break

            print("  ⚠️ Use: s, a, n, p, v ou q.")

    _resumo(verde, amarelo, vermelho, puladas, total)
    input("\n  Pressione ENTER pra voltar ao menu...")


def _salvar(wb):
    """Salva a planilha (sem print pra não poluir)."""
    try:
        wb.save(ARQUIVO_PLANILHA)
    except PermissionError:
        print("  ⚠️ ATENÇÃO: planilha aberta no Excel, não foi possível salvar!")
        print("  Feche o Excel pra continuar salvando.")
    except Exception as e:
        print(f"  ⚠️ Erro ao salvar: {e}")


def _resumo(verde, amarelo, vermelho, puladas, total):
    print("\n" + "═" * 65)
    print("  ✨ RESUMO DA SESSÃO")
    print("═" * 65)
    print(f"  🟢 OK:           {verde}")
    print(f"  🟡 Atenção:      {amarelo}")
    print(f"  🔴 Quebrados:    {vermelho}")
    print(f"  ⏭️  Puladas:      {puladas}")
    print(f"  ──────────────────────")
    print(f"  📊 Processados: {verde + amarelo + vermelho + puladas}/{total}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "=" * 65)
    print("  🔍 VERIFICADOR DE LINKS — OfertasTec")
    print("=" * 65)

    if not Path(ARQUIVO_PLANILHA).exists():
        print(f"\n❌ Planilha '{ARQUIVO_PLANILHA}' não encontrada!")
        input("\nPressione ENTER...")
        return

    print(f"\n📊 Carregando '{ARQUIVO_PLANILHA}'...")
    try:
        wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
    except PermissionError:
        print("\n❌ Planilha ABERTA no Excel! Feche e tente de novo.")
        input("\nPressione ENTER...")
        return
    except Exception as e:
        print(f"\n❌ Erro: {e}")
        input("\nPressione ENTER...")
        return

    if ABA_PLANILHA not in wb.sheetnames:
        print(f"\n❌ Aba '{ABA_PLANILHA}' não encontrada!")
        input("\nPressione ENTER...")
        return

    ws = wb[ABA_PLANILHA]
    produtos = carregar_produtos(ws)

    if not produtos:
        print("\n❌ Nenhum produto com link encontrado.")
        input("\nPressione ENTER...")
        return

    print(f"✅ {len(produtos)} produtos carregados")

    # Loop principal
    while True:
        # recarrega cores atuais (caso tenha mudado)
        for p in produtos:
            p['cor'] = cor_atual(ws, p['row'])

        lista = menu_filtro(produtos)
        if lista is None:
            break
        processar_lista(ws, wb, lista)

    _salvar(wb)
    print("\n💾 Planilha salva. Até a próxima!\n")
    input("Pressione ENTER pra sair...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n👋 Saindo...")
    except Exception as e:
        print(f"\n❌ ERRO inesperado: {e}")
        input("\nPressione ENTER pra sair...")
