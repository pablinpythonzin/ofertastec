"""
═══════════════════════════════════════════════════════════════════════
  PROCESSAR FOTOS — OfertasTec
  Lê o ID do NOME DO ARQUIVO automaticamente (10.webp = produto #10)
═══════════════════════════════════════════════════════════════════════

COMO FUNCIONA:
  Você renomeia as fotos com o ID do produto (10.webp, 11.webp, 15.webp).
  O script lê o número, busca o produto na planilha e renomeia SOZINHO.
  Você só confirma no final. ZERO digitação produto por produto.

COMO USAR:
  1. Roda criar_pastas.py primeiro (cria as pastas)
  2. Baixa fotos e renomeia com o ID (ex: produto #10 → 10.webp)
  3. Joga nas pastas certas: fotos-webp/mouse-com-fio/10.webp
  4. Roda: python processar_fotos.py
  5. Escolhe a categoria
  6. Script mostra o mapeamento e você confirma com 's'

SISTEMA DE CORES (a foto NÃO é processada se a linha estiver):
  🔴 Vermelho = quebrado    → PULA
  🟣 Roxo     = duplicado   → PULA
  (verde, amarelo e sem-cor são processados normalmente)

═══════════════════════════════════════════════════════════════════════
"""

import os
import sys
import unicodedata
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("\n❌ ERRO: 'openpyxl' não instalado. Rode: pip install openpyxl pillow")
    sys.exit(1)

try:
    from PIL import Image
except ImportError:
    print("\n❌ ERRO: 'Pillow' não instalado. Rode: pip install pillow")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════

PASTA_FOTOS_BRUTAS = "fotos-webp"
PASTA_DESTINO = "produtos"
ARQUIVO_PLANILHA = "Produtos_ml.xlsx"
ABA_PLANILHA = "Produtos Renomear"

# Colunas (confirmadas pelo Pablo)
COL_NUM = 2          # B - #
COL_CATEGORIA = 3    # C - Categoria
COL_LINK = 4         # D - Link
COL_NOME = 5         # E - Nome do Produto

# Cores que fazem PULAR a foto
COR_VERMELHO = "FFC00000"
COR_ROXO = "FF7030A0"

# Lista DEFINITIVA: nome-da-pasta : categoria-na-planilha (30 categorias)
CATEGORIAS = {
    "mouse-sem-fio":          "Mouse sem fio",
    "mouse-com-fio":          "Mouse com fio",
    "mouse-gamer-sem-fio":    "Mouse gamer sem fio",
    "mouse-gamer-com-fio":    "Mouse gamer com fio",
    "mouse-vertical":         "Mouse vertical (ergonômico)",
    "teclado-com-fio":        "Teclado com fio",
    "teclado-sem-fio":        "Teclado sem fio",
    "teclado-gamer-com-fio":  "Teclado gamer com fio",
    "teclado-gamer-sem-fio":  "Teclado gamer sem fio",
    "fone-sem-fio":           "Fone sem fio (bluetooth)",
    "fone-com-fio":           "Fone com fio",
    "fone-gamer-sem-fio":     "Fone gamer (headset sem fio)",
    "fone-gamer-com-fio":     "Fone gamer (headset com fio)",
    "caixa-som-bluetooth":    "Caixa de som bluetooth",
    "caixa-som-com-fio":      "Caixa de som com fio",
    "cabo-usbc":              "Cabo USB-C",
    "cabo-lightning":         "Cabo Lightning (iPhone)",
    "power-bank":             "Power bank",
    "suporte-celular-carro":  "Suporte p/ celular (carro)",
    "pelicula-protetora":     "Película protetora",
    "capinha-celular":        "Capinha celular",
    "mousepad-simples":       "Mousepad simples",
    "mousepad-gamer":         "Mousepad gamer (grande)",
    "controle-pc-sem-fio":    "Controle PC (sem fio)",
    "controle-pc-com-fio":    "Controle PC (com fio)",
    "controle-celular":       "Controle para celular",
    "adaptador-tomada":       "Adaptador de tomada",
    "cabo-hdmi":              "Cabo HDMI",
    "pen-drive":              "Pen drive",
    "hd-externo":             "HD externo",
}


# ═══════════════════════════════════════════════════════════════════════
# UTILITÁRIAS
# ═══════════════════════════════════════════════════════════════════════

def remover_acentos(texto):
    if not texto:
        return ""
    nfkd = unicodedata.normalize('NFKD', str(texto))
    return ''.join([c for c in nfkd if not unicodedata.combining(c)])


def gerar_nome_arquivo(nome_produto):
    nome = remover_acentos(nome_produto).lower()
    nome = re.sub(r'[^a-z0-9]+', '-', nome)
    nome = re.sub(r'-+', '-', nome).strip('-')
    return nome


def extrair_id(nome_arquivo):
    """Pega o primeiro número do nome: '10.webp' → 10, 'foto-15.jpg' → 15"""
    m = re.search(r'\d+', nome_arquivo)
    return int(m.group()) if m else None


def converter_e_salvar(origem, destino):
    try:
        img = Image.open(origem)
        if img.mode in ('RGBA', 'LA', 'P'):
            fundo = Image.new('RGB', img.size, (255, 255, 255))
            if img.mode == 'P':
                img = img.convert('RGBA')
            fundo.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
            img = fundo
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        img.save(destino, 'JPEG', quality=85, optimize=True)
        return True, f"{os.path.getsize(destino)/1024:.1f} KB"
    except Exception as e:
        return False, str(e)


def cor_linha(ws, row):
    cell = ws.cell(row=row, column=COL_NOME)
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
        rgb = cell.fill.fgColor.rgb
        return None if rgb in ('00000000', None) else rgb
    return None


def carregar_planilha():
    if not Path(ARQUIVO_PLANILHA).exists():
        print(f"\n❌ Planilha '{ARQUIVO_PLANILHA}' não encontrada!")
        return None
    try:
        wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
    except PermissionError:
        print("\n❌ Planilha ABERTA no Excel! Feche e tente de novo.")
        return None
    ws = wb[ABA_PLANILHA]

    # dados[id] = {nome, cat, cor}
    dados = {}
    for row in range(2, ws.max_row + 1):
        num = ws.cell(row=row, column=COL_NUM).value
        cat = ws.cell(row=row, column=COL_CATEGORIA).value
        nome = ws.cell(row=row, column=COL_NOME).value
        if num and cat and nome and str(nome).strip():
            dados[int(num)] = {
                'nome': str(nome).strip(),
                'cat': str(cat).strip(),
                'cor': cor_linha(ws, row),
            }
    return dados


# ═══════════════════════════════════════════════════════════════════════
# MENU
# ═══════════════════════════════════════════════════════════════════════

def mostrar_menu(dados, pasta_brutas, pasta_destino):
    print("\n" + "=" * 65)
    print("  📁 CATEGORIAS")
    print("=" * 65)
    print(f"\n  {'#':>3} | {'Categoria':28} | {'Fotos':>6} | {'No site':>8}")
    print("  " + "─" * 60)

    opcoes = []
    for i, (slug, nome_pl) in enumerate(CATEGORIAS.items(), 1):
        pasta_cat = pasta_brutas / slug
        exts = ['.webp', '.jpg', '.jpeg', '.png']
        n_fotos = len([f for f in pasta_cat.iterdir()
                       if f.is_file() and f.suffix.lower() in exts]) if pasta_cat.exists() else 0

        produtos_cat = [d for d in dados.values() if d['cat'] == nome_pl]
        com_foto = sum(1 for d in produtos_cat
                       if (pasta_destino / (gerar_nome_arquivo(d['nome']) + '.jpg')).exists())

        status_pasta = "❌ s/pasta" if not pasta_cat.exists() else (str(n_fotos) if n_fotos else "vazia")
        print(f"  [{i:>2}] | {slug:28} | {status_pasta:>6} | {com_foto:>3}/{len(produtos_cat):<3}")
        opcoes.append((slug, nome_pl))

    print(f"\n  [0] Sair")
    while True:
        r = input("\n  Escolha [0 pra sair]: ").strip()
        try:
            n = int(r)
            if n == 0:
                return None
            if 1 <= n <= len(opcoes):
                return opcoes[n - 1]
        except ValueError:
            pass
        print("  ⚠️ Número inválido.")


# ═══════════════════════════════════════════════════════════════════════
# PROCESSAR CATEGORIA (modo automático por ID)
# ═══════════════════════════════════════════════════════════════════════

def processar(slug, nome_pl, dados, pasta_brutas, pasta_destino):
    pasta_cat = pasta_brutas / slug
    if not pasta_cat.exists():
        print(f"\n  ❌ Pasta '{slug}' não existe! Rode criar_pastas.py")
        input("\n  ENTER..."); return

    exts = ['.webp', '.jpg', '.jpeg', '.png']
    fotos = sorted([f for f in pasta_cat.iterdir()
                    if f.is_file() and f.suffix.lower() in exts])
    if not fotos:
        print(f"\n  ❌ Nenhuma foto em '{slug}/'")
        input("\n  ENTER..."); return

    print("\n" + "=" * 65)
    print(f"  📁 {slug}  ({nome_pl})")
    print("=" * 65)

    # montar mapeamento
    plano = []      # (foto, id, nome_arquivo)
    problemas = []  # (foto, motivo)

    for foto in fotos:
        pid = extrair_id(foto.stem)
        if pid is None:
            problemas.append((foto.name, "sem número no nome"))
        elif pid not in dados:
            problemas.append((foto.name, f"produto #{pid} não existe na planilha"))
        elif dados[pid]['cat'] != nome_pl:
            problemas.append((foto.name, f"#{pid} é '{dados[pid]['cat']}', não '{nome_pl}'"))
        elif dados[pid]['cor'] == COR_VERMELHO:
            problemas.append((foto.name, f"#{pid} está VERMELHO (quebrado)"))
        elif dados[pid]['cor'] == COR_ROXO:
            problemas.append((foto.name, f"#{pid} está ROXO (duplicado)"))
        else:
            nome_arq = gerar_nome_arquivo(dados[pid]['nome']) + '.jpg'
            plano.append((foto, pid, nome_arq))

    # mostrar plano
    if plano:
        print(f"\n  ✅ {len(plano)} foto(s) prontas pra processar:\n")
        for foto, pid, nome_arq in sorted(plano, key=lambda x: x[1]):
            print(f"     #{pid:>3} | {foto.name:18} → {nome_arq}")

    if problemas:
        print(f"\n  ⚠️ {len(problemas)} foto(s) com problema (serão ignoradas):\n")
        for nome, motivo in problemas:
            print(f"     {nome:18} → {motivo}")

    if not plano:
        print("\n  ❌ Nada pra processar.")
        input("\n  ENTER..."); return

    print()
    if input("  ✅ Confirma processar? (s/n): ").strip().lower() != 's':
        print("  Cancelado."); input("\n  ENTER..."); return

    # processar
    print("\n  ⚙️  Processando...\n")
    ok_count = err = pulado = 0
    for foto, pid, nome_arq in sorted(plano, key=lambda x: x[1]):
        destino = pasta_destino / nome_arq
        if destino.exists():
            print(f"  ⏭️  #{pid:>3} | já existe: {nome_arq}")
            pulado += 1
            continue
        ok, info = converter_e_salvar(str(foto), str(destino))
        if ok:
            print(f"  ✅ #{pid:>3} | {nome_arq} ({info})")
            ok_count += 1
        else:
            print(f"  ❌ #{pid:>3} | erro: {info}")
            err += 1

    print(f"\n  ✨ {ok_count} processadas | {pulado} já existiam | {err} erros")
    input("\n  ENTER pra voltar ao menu...")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "=" * 65)
    print("  🚀 PROCESSAR FOTOS — OfertasTec (auto por ID)")
    print("=" * 65)

    pasta_atual = Path.cwd()
    pasta_brutas = pasta_atual / PASTA_FOTOS_BRUTAS
    pasta_destino = pasta_atual / PASTA_DESTINO

    if not pasta_brutas.exists():
        print(f"\n❌ Pasta '{PASTA_FOTOS_BRUTAS}/' não existe! Rode criar_pastas.py")
        input("\nENTER..."); return
    pasta_destino.mkdir(exist_ok=True)

    print("\n📊 Carregando planilha...")
    dados = carregar_planilha()
    if dados is None:
        input("\nENTER..."); return
    print(f"✅ {len(dados)} produtos carregados")

    while True:
        escolha = mostrar_menu(dados, pasta_brutas, pasta_destino)
        if escolha is None:
            break
        slug, nome_pl = escolha
        processar(slug, nome_pl, dados, pasta_brutas, pasta_destino)

    print("\n📁 Fotos em: produtos/")
    print("📝 Próximo: peça ao Claude pra atualizar o index.html\n")
    input("ENTER pra sair...")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n👋 Saindo...")
    except Exception as e:
        print(f"\n❌ ERRO: {e}")
        input("\nENTER pra sair...")
